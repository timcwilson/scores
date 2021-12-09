#!/usr/bin/env python3
import copy

import requests
import wx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

comp_scores = {}
STATE_START = 0
STATE_RXD_SCORES = 1
STATE_RXD_SSHEET = 2
STATE_PROCESS = 3
state = STATE_START
ss_file = ""


class Scores(wx.Frame):
    def __init__(self, *args, **kwds):
        # begin wxGlade: Scores.__init__
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_FRAME_STYLE
        wx.Frame.__init__(self, *args, **kwds)
        self.SetSize((400, 300))
        self.SetTitle("SCA Live Scores Team Processing")

        self.panel_1 = wx.Panel(self, wx.ID_ANY)

        sizer_1 = wx.BoxSizer(wx.VERTICAL)

        sizer_1.Add((20, 20), 0, 0, 0)

        lblHeading = wx.StaticText(
            self.panel_1, wx.ID_ANY, "SCA Live Scores Team Processing"
        )
        lblHeading.SetFont(
            wx.Font(
                20,
                wx.FONTFAMILY_DEFAULT,
                wx.FONTSTYLE_NORMAL,
                wx.FONTWEIGHT_BOLD,
                0,
                "",
            )
        )
        sizer_1.Add(lblHeading, 0, wx.ALIGN_CENTER_HORIZONTAL, 0)

        sizer_1.Add((20, 20), 0, 0, 0)

        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_1.Add(sizer_2, 1, wx.EXPAND, 0)

        sizer_2.Add((20, 20), 0, 0, 0)

        lblLiveScoresURL = wx.StaticText(self.panel_1, wx.ID_ANY, "SCA Live Scores URL")
        lblLiveScoresURL.SetMinSize((140, 16))
        sizer_2.Add(lblLiveScoresURL, 0, 0, 0)

        sizer_2.Add((10, 20), 0, 0, 0)

        self.txtLiveScoresURL = wx.TextCtrl(
            self.panel_1, wx.ID_ANY, "", style=wx.TE_PROCESS_ENTER | wx.TE_PROCESS_TAB
        )
        self.txtLiveScoresURL.SetMinSize((205, 22))
        sizer_2.Add(self.txtLiveScoresURL, 0, 0, 0)

        sizer_2.Add((20, 20), 0, 0, 0)

        sizer_3 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_1.Add(sizer_3, 1, wx.EXPAND, 0)

        sizer_3.Add((20, 20), 0, 0, 0)

        self.lblTeamSpreadsheet = wx.StaticText(
            self.panel_1, wx.ID_ANY, "Team Spreadsheet"
        )
        self.lblTeamSpreadsheet.SetMinSize((140, 16))
        sizer_3.Add(self.lblTeamSpreadsheet, 0, 0, 0)

        sizer_3.Add((10, 20), 0, 0, 0)

        self.txtTeamSpreadsheet = wx.TextCtrl(
            self.panel_1, wx.ID_ANY, "", style=wx.TE_READONLY
        )
        self.txtTeamSpreadsheet.SetMinSize((205, 22))
        sizer_3.Add(self.txtTeamSpreadsheet, 0, 0, 0)

        sizer_3.Add((20, 20), 0, 0, 0)

        sizer_4 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_1.Add(sizer_4, 1, wx.EXPAND, 0)

        sizer_4.Add((20, 20), 0, 0, 0)

        self.lblSpreadsheetCol = wx.StaticText(
            self.panel_1, wx.ID_ANY, "Spreadsheet Column"
        )
        self.lblSpreadsheetCol.SetMinSize((140, 16))
        sizer_4.Add(self.lblSpreadsheetCol, 0, 0, 0)

        sizer_4.Add((10, 20), 0, 0, 0)

        self.txtSpreadsheetColumn = wx.TextCtrl(self.panel_1, wx.ID_ANY, "")
        self.txtSpreadsheetColumn.SetMinSize((205, 22))
        sizer_4.Add(self.txtSpreadsheetColumn, 0, 0, 0)

        sizer_4.Add((20, 20), 0, 0, 0)

        sizer_5 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_1.Add(sizer_5, 1, wx.EXPAND, 0)

        sizer_5.Add((20, 20), 0, 0, 0)

        self.btnMulti = wx.Button(self.panel_1, wx.ID_ANY, "Check Scores URL")
        self.btnMulti.SetMinSize((140, 21))
        sizer_5.Add(self.btnMulti, 0, 0, 0)

        sizer_5.Add((10, 20), 0, 0, 0)

        self.btnProcess = wx.Button(self.panel_1, wx.ID_ANY, "Process Scores")
        self.btnProcess.SetMinSize((205, 21))
        sizer_5.Add(self.btnProcess, 0, wx.ALL, 0)

        sizer_5.Add((20, 20), 0, 0, 0)

        self.panel_1.SetSizer(sizer_1)

        self.Layout()
        self.setState(STATE_START)
        self.Bind(wx.EVT_TEXT_ENTER, self.tstURL, self.txtLiveScoresURL)
        self.Bind(wx.EVT_BUTTON, self.btnPressMulti, self.btnMulti)
        self.Bind(wx.EVT_BUTTON, self.btnPressProcess, self.btnProcess)

    def tstURL(self, event):
        global comp_scores
        comp_scores = {}
        url = self.txtLiveScoresURL.GetValue()
        try:
            html_text = requests.get(url).text
        except:
            wx.MessageBox("Invalid URL. No scores found. Please try again.")
            return

        soup = BeautifulSoup(html_text, "html.parser")
        competitor_list = soup.find_all("tbody")
        for tbl in competitor_list:
            competitor_in_table = tbl.find_all("tr")
            for competitor in competitor_in_table:
                results = competitor.find_all("td")
                comp_scores[results[0].text] = results[4].text

        self.setState(STATE_RXD_SCORES)

        event.Skip()

    def btnPressMulti(self, event):
        global state, comp_scores, ss_file

        if state == STATE_RXD_SCORES:
            dir_name = ""
            dlg = wx.FileDialog(
                self, "Choose Team Spreadsheet", dir_name, "", "*.xlsm", wx.FD_OPEN
            )
            if dlg.ShowModal() == wx.ID_OK:
                ss_file = f"{dlg.GetDirectory()}/{dlg.GetFilename()}"
                try:
                    wb = load_workbook(filename=ss_file, keep_vba=True)
                    ws = wb.active
                    wb.close()
                    self.txtTeamSpreadsheet.SetValue(ss_file)
                    self.setState(STATE_RXD_SSHEET)
                except:
                    wx.MessageBox(
                        "Error opening spreadsheet. Please make sure it is not open in Excel."
                    )
                    return
            else:
                pass
        elif state == STATE_RXD_SSHEET:
            wb = load_workbook(filename=ss_file, keep_vba=True)
            ws = wb.active
            tmp_scores = copy.deepcopy(comp_scores)
            i = 4
            while i < 1000:
                try:
                    key = ws[f"C{i}"].value + ", " + ws[f"B{i}"].value
                    if key in tmp_scores:
                        tmp_scores.pop(key, 0)
                except:
                    pass
                i += 1

            if len(tmp_scores) > 0:
                print(tmp_scores)
            self.setState(STATE_PROCESS)
        event.Skip()

    def btnPressProcess(self, event):  # wxGlade: Scores.<event_handler>
        global state, comp_scores, ss_file

        if state == STATE_PROCESS:
            col = self.txtSpreadsheetColumn.GetValue()
            wb = load_workbook(filename=ss_file, keep_vba=True)
            ws = wb.active
            i = 4
            while i < 1000:
                try:
                    key = ws[f"C{i}"].value + ", " + ws[f"B{i}"].value
                    if key in comp_scores:
                        ws[f"{col}{i}"].value = int(comp_scores[key])
                except:
                    pass
                i += 1

            wb.save(ss_file)
            wx.MessageBox("Scores processed. Please check spreadsheet.")
            self.txtSpreadsheetColumn.SetValue("")
            self.txtLiveScoresURL.SetValue("")
            self.txtTeamSpreadsheet.SetValue("")
            self.setState(STATE_START)
        else:
            print("Error process incorrect state")
        event.Skip()

    def setState(self, newState):
        global state
        if newState == STATE_START:
            self.txtLiveScoresURL.SetValue("")
            self.lblTeamSpreadsheet.Hide()
            self.txtTeamSpreadsheet.Hide()
            self.lblSpreadsheetCol.Hide()
            self.txtSpreadsheetColumn.Hide()
            self.btnMulti.Hide()
            self.btnProcess.Hide()
        elif newState == STATE_RXD_SCORES:
            self.lblTeamSpreadsheet.Hide()
            self.txtTeamSpreadsheet.Hide()
            self.lblSpreadsheetCol.Hide()
            self.txtSpreadsheetColumn.Hide()
            self.btnMulti.SetLabelText("Open Spreadsheet")
            self.btnMulti.Show()
            self.btnProcess.Hide()
        elif newState == STATE_RXD_SSHEET:
            self.lblTeamSpreadsheet.Show()
            self.txtTeamSpreadsheet.Show()
            self.lblSpreadsheetCol.Hide()
            self.txtSpreadsheetColumn.Hide()
            self.btnMulti.SetLabelText("Test Results")
            self.btnMulti.Show()
            self.btnProcess.Hide()
        elif newState == STATE_PROCESS:
            self.lblTeamSpreadsheet.Show()
            self.txtTeamSpreadsheet.Show()
            self.lblSpreadsheetCol.Show()
            self.txtSpreadsheetColumn.Show()
            self.btnMulti.SetLabelText("Update Spreadsheet")
            self.btnMulti.Show()
            self.btnProcess.Show()
        else:
            print("State Error")
            return
        state = newState


class MyAppScores(wx.App):
    def OnInit(self):
        self.frame = Scores(None, wx.ID_ANY, "")
        self.SetTopWindow(self.frame)
        self.frame.Show()
        return True


if __name__ == "__main__":
    appScores = MyAppScores(0)
    wx.MessageBox("Copy and paste the SCA Lives Scores URL into the text box.")
    appScores.MainLoop()
