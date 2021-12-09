#!/usr/bin/python3

import click
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


@click.command()
@click.option("--URL", prompt="URL", help="Provide URL of shoot")
@click.option(
    "--col", prompt="Column for scores", help="Provide column to store scores"
)
def score(url, col):
    wb = load_workbook(filename="2019 Vic State Team Sheet -1.xlsm", keep_vba=True)
    ws = wb.active
    click.echo(f"URL: {url}")
    click.echo(f"Column: {col}")
    comp_scores = {}
    html_text = requests.get(url).text
    soup = BeautifulSoup(html_text, "html.parser")
    competitor_list = soup.find_all("tbody")
    for tbl in competitor_list:
        competitor_in_table = tbl.find_all("tr")
        for competitor in competitor_in_table:
            results = competitor.find_all("td")
            comp_scores[results[0].text] = results[4].text

    i = 4
    while i < 1000:
        try:
            key = ws[f"C{i}"].value + ", " + ws[f"B{i}"].value
            print(key)
            if key in comp_scores:
                print(comp_scores[key])
                ws[f"{col}{i}"].value = int(comp_scores[key])
        except:
            pass
        i += 1

    wb.save("2019 Vic State Team Sheet -1.xlsm")


if __name__ == "__main__":
    score()
